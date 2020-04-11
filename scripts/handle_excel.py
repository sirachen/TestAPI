# coding=utf-8
'''
------------------------------------
 @Auth : want
 @Time : 2020/3/24 下午5:12
 @File : handle_excel.py
 @IDE  : PyCharm
------------------------------------
'''


from openpyxl import load_workbook

import os

from scripts.path_constants import DATA_CASES


class HandleExcel:

    def __init__(self, filename=DATA_CASES, sheetname=None):
        self.filename = filename
        self.sheetname = sheetname

    def get_cases(self):
        '''
        获取所有测试用例
        '''
        work_book = load_workbook(self.filename)

        if self.sheetname is None:
            worksheet = work_book.active
        else:
            worksheet = work_book[self.sheetname]

        head_tuple = tuple(worksheet.iter_rows(max_row=1, values_only=True))[0]

        value_list = []
        for tuple_v in tuple(worksheet.iter_rows(min_row=2, values_only=True)):
            value_list.append(dict(zip(head_tuple, tuple_v)))

        return value_list

    def get_case(self, row):

        return self.get_cases()[row-1]

    def write_cell(self, row, column, actual, result):

        write_work_book = load_workbook(self.filename)

        if self.sheetname is None:
            write_worksheet = write_work_book.active
        else:
            write_worksheet = write_work_book[self.sheetname]

        if isinstance(row, int) and (2 <= row <= write_worksheet.max_row):
            write_worksheet.cell(row=row, column=column, value=actual)
            write_worksheet.cell(row=row, column=column, value=result)
            write_work_book.save(self.filename)
            write_work_book.close()
        else:
            print('行号有误,应为大于2的整数')

    # def write_cell(self, row, column, expected, actual, result):
    #
    #     write_work_book = load_workbook(self.filename)
    #
    #     if self.sheetname is None:
    #         write_worksheet = write_work_book.active
    #     else:
    #         write_worksheet = write_work_book[self.sheetname]
    #
    #     if isinstance(row, int) and (2 <= row <= write_worksheet.max_row):
    #         write_worksheet.cell(row=row, column=column, value=expected)
    #         write_worksheet.cell(row=row, column=column+1, value=actual)
    #         write_worksheet.cell(row=row, column=column+2, value=result)
    #         write_work_book.save(self.filename)
    #         write_work_book.close()
    #     else:
    #         print('行号有误,应为大于2的整数')


if __name__ == '__main__':
    handle_excel = HandleExcel()
    print(handle_excel.get_cases())
    print(handle_excel.get_case(row=1))
    # handle_excel.write_cell(5, 6, 'a', 'b')